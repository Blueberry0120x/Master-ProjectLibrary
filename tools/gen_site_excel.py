"""Generate / update SiteMasterList.xlsx from site_data.json files.

First run  → creates the workbook fresh with openpyxl (full formatting).
Subsequent → opens existing file via COM (win32com) and updates only data
             cells A2:R{n+1}.  The Notes column (S) and any manual edits
             the user made in later columns are never touched.
"""
from __future__ import annotations
import json
import os
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parent.parent
SITES_DIR = REPO / "data" / "sites"
OUT = REPO / "data" / "SiteMasterList.xlsx"

# ── colour palette ──────────────────────────────────────────────────────────
C_HEADER   = "0F4C81"
C_SD       = "D6E4F0"
C_OC       = "FFF3CD"
C_WA       = "D5F0D5"
C_YES      = "C6EFCE"
C_NO       = "FFCCCC"
C_FILE_OK  = "C6EFCE"
C_FILE_NO  = "FFCCCC"

DROPBOX = {
    "ca-2921-el-cajon":  "2921-2923 El Cajon Blvd/",
    "ca-4335-euclid":    "4335 Euclid Ave San Diego, CA 92115/",
    "ca-1905-rohn":      "1905 Rohn Rd Escondido, CA 92025/",
    "ca-12652-laux":     "12652 Laux Ave, Garden Grove, CA 92840/",
    "ca-5251-palmyra":   "5251 Palmyra Ave, San Diego CA 92117/",
    "ca-9362-angwin":    "9362 Angwin Pl San Diego, CA 92123 United States/",
    "wa-405-126th":      "405 SW 126th St Seattle, WA 98146/",
}
CITY_COUNTY = {
    "San Diego": "San Diego", "Escondido": "San Diego",
    "Garden Grove": "Orange", "Kirkland": "King, WA",
    "Renton": "King, WA", "Burien": "King, WA",
}
SITE_ORDER = [
    "ca-2921-el-cajon", "ca-3063-cabrillo-mesa", "ca-4335-euclid",
    "ca-4876-cannington", "ca-5251-palmyra", "ca-9362-angwin",
    "ca-1905-rohn", "ca-11001-westminster", "ca-12652-laux",
    "wa-10404-kirkland", "wa-12843-175th", "wa-405-126th",
]


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border() -> Border:
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)


INDEX_STATUS: dict[str, str] = {}


def _load_index() -> None:
    fp = SITES_DIR / "index.json"
    if not fp.exists():
        return
    data = json.loads(fp.read_text(encoding="utf-8"))
    for entry in data.get("sites", []):
        raw = entry["id"].lower().replace("_", "-")
        INDEX_STATUS[raw] = entry.get("status", "skeleton")


def load_sites() -> list[dict]:
    _load_index()
    rows: list[dict] = []
    for site_id in SITE_ORDER:
        folder = SITES_DIR / site_id
        fp = folder / "site_data.json"
        if not fp.exists():
            continue
        d = json.loads(fp.read_text(encoding="utf-8"))

        addr_raw = d.get("address", "")
        if isinstance(addr_raw, dict):
            full_addr = addr_raw.get("full", "")
            city = addr_raw.get("city", "")
        else:
            full_addr = str(addr_raw)
            parts = full_addr.split(",")
            city = parts[1].strip() if len(parts) > 1 else ""

        county = CITY_COUNTY.get(city, "")
        state = site_id[:2].upper()

        apn = d.get("apn", "") or ""
        if isinstance(apn, list):
            apn = ", ".join(apn)

        zoning = ""
        if isinstance(d.get("site"), dict):
            zoning = d["site"].get("zoning", "")
        zoning = zoning or d.get("zoning", "") or ""

        files_in = [f.name for f in folder.iterdir() if f.name != "site_data.json"]
        pdfs = [f for f in files_in if f.lower().endswith(".pdf")]
        dwgs = [f for f in files_in if f.lower().endswith(".dwg")]
        parcel_maps = [f for f in pdfs
                       if "parcelmap" in f.lower() or "parcel_map" in f.lower()]
        surveys = [f for f in pdfs + dwgs
                   if any(k in f.lower() for k in ("survey", "topo", "rohn", "sv-"))]

        idx_st = INDEX_STATUS.get(site_id, "skeleton")
        proj = d.get("project", "")
        has_proj = isinstance(proj, dict) and bool(proj)

        rows.append({
            "site_id":    site_id,
            "address":    full_addr,
            "apn":        apn,
            "county":     county,
            "state":      state,
            "zoning":     zoning,
            "dropbox":    DROPBOX.get(site_id, ""),
            "parcel_map": "; ".join(parcel_maps),
            "survey":     "; ".join(surveys),
            "s1": "Y" if has_proj else "",
            "s2": "Y" if apn else "",
            "s3": "Y" if zoning else "",
            "s7": "Y" if parcel_maps else "",
            "s5": "Y" if idx_st in ("complete", "partial") else "",
            "s4": "",
            "s6": "",
            "s8": "Y" if surveys else "",
        })
    return rows


# ── row values in column order ────────────────────────────────────────────────
def _row_values(i: int, r: dict) -> list:
    return [
        i, r["site_id"], r["address"], r["apn"],
        r["county"], r["state"], r["zoning"], r["dropbox"],
        r["parcel_map"], r["survey"],
        r["s1"], r["s2"], r["s3"], r["s7"],
        r["s5"], r["s4"], r["s6"], r["s8"],
        "",   # Notes — never overwritten after first creation
    ]


STEP_COLS  = set(range(11, 19))   # cols 11-18 (1-based) = step columns
FILE_COLS  = {9, 10}              # parcel map, survey


# ══════════════════════════════════════════════════════════════════════════════
#  CREATE (openpyxl — first run)
# ══════════════════════════════════════════════════════════════════════════════

HEADERS = [
    "#", "Site ID", "Full Address", "APN(s)", "County", "St",
    "Zoning", "Dropbox Folder",
    "Parcel Map (repo)", "Survey (repo)",
    "S1 Proj Info", "S2 APN", "S3 Zoning",
    "S7 Parcel Map", "S5 Site Visit",
    "S4 Photos", "S6 Survey Req", "S8 Survey Rcvd",
    "Notes",
]
COL_WIDTHS = [4, 22, 42, 22, 14, 4, 14, 40, 34, 34,
              10, 8, 8, 12, 12, 8, 11, 12, 20]


def _county_color(r: dict) -> str:
    if r["county"] == "Orange":
        return C_OC
    if r["state"] == "WA":
        return C_WA
    return C_SD


def _cell_fill(col: int, val: str, row_color: str) -> PatternFill:
    if col in STEP_COLS:
        return _fill(C_YES if val == "Y" else C_NO)
    if col in FILE_COLS:
        return _fill(C_FILE_OK if val else C_FILE_NO)
    return _fill(row_color)


def create_workbook(rows: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Site Master List"

    # header row
    ws.append(HEADERS)
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=ci)
        cell.font = hdr_font
        cell.fill = _fill(C_HEADER)
        cell.alignment = hdr_align
        cell.border = _border()
    ws.row_dimensions[1].height = 36

    # data rows
    for row_num, r in enumerate(rows, start=2):
        rc = _county_color(r)
        values = _row_values(row_num - 1, r)
        ws.append(values)
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=ci)
            cell.border = _border()
            cell.alignment = Alignment(vertical="top", wrap_text=(ci in (3, 8, 9, 10)))
            cell.fill = _cell_fill(ci, str(val) if val else "", rc)
        ws.row_dimensions[row_num].height = 40

    # column widths
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # legend
    lr = len(rows) + 3
    ws.cell(row=lr, column=1, value="LEGEND").font = Font(bold=True)
    for i, (color, label) in enumerate([
        (C_SD, "San Diego County (CA)"), (C_OC, "Orange County (CA)"),
        (C_WA, "Washington (WA)"),       (C_YES, "Step: Complete (Y)"),
        (C_NO, "Step: Not done"),         (C_FILE_OK, "File: Present in repo"),
        (C_FILE_NO, "File: Missing"),
    ], 1):
        r = lr + i
        ws.cell(row=r, column=1).fill = _fill(color)
        ws.cell(row=r, column=1).border = _border()
        ws.cell(row=r, column=2, value=label)

    ws.freeze_panes = "C2"
    wb.save(OUT)
    print(f"Created: {OUT}")


# ══════════════════════════════════════════════════════════════════════════════
#  UPDATE via COM (subsequent runs — preserves Notes + manual edits)
# ══════════════════════════════════════════════════════════════════════════════

#  Column map: 1-based Excel column → row dict key.
#  Notes column (19 / S) is intentionally absent — never overwritten.
_COL_KEY: dict[int, str] = {
    1: "#", 2: "site_id", 3: "address", 4: "apn",
    5: "county", 6: "state", 7: "zoning", 8: "dropbox",
    9: "parcel_map", 10: "survey",
    11: "s1", 12: "s2", 13: "s3", 14: "s7",
    15: "s5", 16: "s4", 17: "s6", 18: "s8",
}

# RGB hex → Excel interior.Color long int (BGR order, 0xBBGGRR)
def _hex_to_long(h: str) -> int:
    h = h.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return b | (g << 8) | (r << 16)   # Excel ColorIndex is 0xRRGGBB as int


_COM_FILL: dict[str, int] = {
    C_YES:     _hex_to_long(C_YES),
    C_NO:      _hex_to_long(C_NO),
    C_FILE_OK: _hex_to_long(C_FILE_OK),
    C_FILE_NO: _hex_to_long(C_FILE_NO),
    C_SD:      _hex_to_long(C_SD),
    C_OC:      _hex_to_long(C_OC),
    C_WA:      _hex_to_long(C_WA),
}


def update_via_com(rows: list[dict]) -> None:
    """Open existing workbook via COM and update data cells A2:R{n+1} only."""
    import win32com.client as win32

    xl = win32.Dispatch("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False

    wb = xl.Workbooks.Open(str(OUT.resolve()))
    ws = wb.Sheets("Site Master List")

    for row_num, r in enumerate(rows, start=2):
        rc = _county_color(r)
        values = _row_values(row_num - 1, r)
        for ci in range(1, 19):   # cols 1-18; skip 19 (Notes)
            cell = ws.Cells(row_num, ci)
            val = values[ci - 1]
            cell.Value = val

            # background colour
            val_str = str(val) if val else ""
            if ci in STEP_COLS:
                color_key = C_YES if val_str == "Y" else C_NO
            elif ci in FILE_COLS:
                color_key = C_FILE_OK if val_str else C_FILE_NO
            else:
                color_key = rc
            cell.Interior.Color = _COM_FILL[color_key]

    wb.Save()
    wb.Close(False)
    xl.Quit()
    print(f"Updated (COM): {OUT}")


# ══════════════════════════════════════════════════════════════════════════════
#  Entry point
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    rows = load_sites()
    if OUT.exists():
        print(f"File exists — updating in place via COM ({len(rows)} sites)…")
        update_via_com(rows)
    else:
        print(f"File not found — creating fresh ({len(rows)} sites)…")
        create_workbook(rows)
    print("Done.")
